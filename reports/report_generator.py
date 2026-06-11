"""Report generation using ReportLab and OpenPyXL."""

from datetime import date, datetime
from decimal import Decimal
from typing import Tuple

from sqlalchemy.orm import Session

from repositories.appointment_repository import AppointmentRepository
from repositories.billing_repository import BillingRepository
from repositories.consultation_repository import ConsultationRepository
from repositories.medicine_repository import MedicineRepository
from repositories.patient_repository import PatientRepository
from repositories.philhealth_repository import PhilHealthTransactionRepository
from reports.pdf_generator import PDFGenerator
from utils.helpers import format_currency
from utils.validators import parse_date


class ReportGenerator:
    def __init__(self, session: Session, settings_service=None, activity_service=None) -> None:
        self.session          = session
        self.activity_service = activity_service
        self.patient_repo     = PatientRepository(session)
        self.consultation_repo = ConsultationRepository(session)
        self.medicine_repo    = MedicineRepository(session)
        self.billing_repo     = BillingRepository(session)
        self.philhealth_repo  = PhilHealthTransactionRepository(session)
        self.settings_service = settings_service

    def _clinic_name(self) -> str:
        if self.settings_service:
            s = self.settings_service.get_settings()
            return (s.clinic_name or "Clinic") if s else "Clinic"
        return "Clinic"

    def generate(
        self, report_type: str, fmt: str, output_path: str,
        start_date: str = "", end_date: str = "",
    ) -> Tuple[bool, str]:
        try:
            if report_type in ("daily_income", "monthly_income", "yearly_income"):
                ok, msg = self._generate_income_report(
                    report_type, fmt, output_path, start_date, end_date
                )
            else:
                data = self._get_data(report_type, start_date, end_date)
                if fmt == "pdf":
                    ok, msg = self._export_pdf(report_type, data, output_path)
                else:
                    ok, msg = self._export_excel(report_type, data, output_path)
            if ok and self.activity_service:
                self.activity_service.log(
                    "EXPORT", "Reports",
                    f"Exported {report_type} report as {fmt.upper()} to {output_path}",
                )
            return ok, msg
        except Exception as exc:
            return False, f"Report generation failed: {exc}"

    # ── Income reports (PDF only for now, but Excel too) ──────────────────────

    def _generate_income_report(
        self, report_type: str, fmt: str, output_path: str,
        start_date: str, end_date: str,
    ) -> Tuple[bool, str]:
        today  = date.today()
        start_d = parse_date(start_date) or today.replace(day=1)
        end_d   = parse_date(end_date)   or today

        if report_type == "daily_income":
            title  = "Daily Income Report"
            period = f"{start_d.strftime('%B %d, %Y')} – {end_d.strftime('%B %d, %Y')}"
        elif report_type == "monthly_income":
            title  = "Monthly Income Report"
            period = f"{start_d.strftime('%B %Y')} – {end_d.strftime('%B %Y')}"
        else:
            title  = "Yearly Income Report"
            period = f"{start_d.year} – {end_d.year}"

        billings = self.billing_repo.get_by_date_range(start_d, end_d)
        summary  = self.billing_repo.get_income_summary(start_d, end_d)

        rows = []
        for b in billings:
            patient_name = b.patient.full_name if b.patient else f"Patient #{b.patient_id}"
            rows.append({
                "date":       b.created_at.strftime("%Y-%m-%d"),
                "billing_no": b.billing_number,
                "patient":    patient_name,
                "total":      float(b.total_amount),
                "paid":       float(b.amount_paid),
                "balance":    float(b.balance),
                "status":     b.payment_status,
            })

        if fmt == "pdf":
            PDFGenerator.generate_income_report(
                output_path  = output_path,
                clinic_name  = self._clinic_name(),
                report_title = title,
                period_label = period,
                rows         = rows,
                summary      = summary,
            )
            return True, f"PDF report saved to {output_path}"

        return self._export_income_excel(output_path, title, period, rows, summary)

    def _export_income_excel(
        self, path: str, title: str, period: str,
        rows: list, summary: dict,
    ) -> Tuple[bool, str]:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Income Report"

        hdr_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        hdr_font = Font(color="FFFFFF", bold=True)
        sum_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        sum_font = Font(bold=True)

        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Period: {period}"
        ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        ws.append([])
        ws.append(["Summary"])
        summary_row = [
            f"Total Billings: ₱{summary['total_billed']:,.2f}",
            f"Collected: ₱{summary['total_collected']:,.2f}",
            f"Balance: ₱{summary['total_balance']:,.2f}",
            f"Count: {summary['count']}",
        ]
        ws.append(summary_row)
        for cell in ws[ws.max_row]:
            cell.fill = sum_fill
            cell.font = sum_font

        ws.append([])
        headers = ["Date", "Bill No.", "Patient", "Total (₱)", "Paid (₱)", "Balance (₱)", "Status"]
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.fill = hdr_fill
            cell.font = hdr_font

        for r in rows:
            ws.append([
                r["date"], r["billing_no"], r["patient"],
                r["total"], r["paid"], r["balance"], r["status"],
            ])

        wb.save(path)
        return True, f"Excel report saved to {path}"

    # ── Generic reports ───────────────────────────────────────────────────────

    def _get_data(self, report_type: str, start: str, end: str) -> dict:
        start_d = parse_date(start) or date.today().replace(day=1)
        end_d   = parse_date(end)   or date.today()

        if report_type == "patients":
            return {"rows": self.patient_repo.search(""),
                    "headers": ["ID", "Patient No.", "Name", "Contact", "PhilHealth"]}
        if report_type == "inventory":
            return {"rows": self.medicine_repo.search(""),
                    "headers": ["ID", "Generic Name", "Brand", "Stock", "Price", "Expiry"]}
        if report_type == "low_stock":
            return {"rows": self.medicine_repo.get_low_stock(),
                    "headers": ["ID", "Generic Name", "Stock", "Reorder Level"]}
        if report_type == "expiring":
            return {"rows": self.medicine_repo.get_expiring(),
                    "headers": ["ID", "Generic Name", "Stock", "Expiry Date"]}
        if report_type == "consultations":
            return {"rows": self.consultation_repo.get_all(),
                    "headers": ["ID", "Patient ID", "Date", "Diagnosis", "Status"]}
        if report_type == "philhealth":
            return {"rows": self.philhealth_repo.get_all(),
                    "headers": ["ID", "Patient ID", "Deduction", "Balance", "Date"]}
        if report_type == "billing":
            billings = self.billing_repo.get_by_date_range(start_d, end_d)
            return {"rows": billings,
                    "headers": ["Bill No.", "Patient ID", "Total", "Paid", "Balance", "Status"]}
        return {"rows": [], "headers": []}

    def _row_values(self, report_type: str, row) -> list:
        if report_type == "patients":
            return [row.id, row.patient_number, row.full_name,
                    row.contact_number or "", row.philhealth_number or ""]
        if report_type in ("inventory", "low_stock", "expiring"):
            vals = [row.id, row.generic_name, row.brand_name or "", row.stock_quantity]
            if report_type == "inventory":
                vals.extend([float(row.selling_price), str(row.expiration_date or "")])
            elif report_type == "low_stock":
                vals.append(row.reorder_level)
            else:
                vals.append(str(row.expiration_date or ""))
            return vals
        if report_type == "consultations":
            return [row.id, row.patient_id, str(row.consultation_date)[:16],
                    (row.diagnosis or "")[:50], row.status]
        if report_type == "philhealth":
            return [row.id, row.patient_id, float(row.philhealth_deduction),
                    float(row.patient_balance), str(row.transaction_date)[:10]]
        if report_type == "billing":
            return [row.billing_number, row.patient_id, float(row.total_amount),
                    float(row.amount_paid), float(row.balance), row.payment_status]
        return []

    def _export_pdf(self, report_type: str, data: dict, path: str) -> Tuple[bool, str]:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        doc = SimpleDocTemplate(path, pagesize=landscape(letter))
        elements = []
        styles = getSampleStyleSheet()
        title = report_type.replace("_", " ").title()
        elements.append(Paragraph(f"{self._clinic_name()} — {title}", styles["Title"]))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
        elements.append(Spacer(1, 20))

        table_data = [data["headers"]]
        for row in data["rows"]:
            table_data.append([str(v) for v in self._row_values(report_type, row)])

        if len(table_data) > 1:
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
                ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",   (0, 0), (-1, -1), 9),
                ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FA")]),
            ]))
            elements.append(table)
        else:
            elements.append(Paragraph("No data available.", styles["Normal"]))

        doc.build(elements)
        return True, f"PDF report saved to {path}"

    def _export_excel(self, report_type: str, data: dict, path: str) -> Tuple[bool, str]:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = report_type[:31]

        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(data["headers"], 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, row in enumerate(data["rows"], 2):
            for col_idx, val in enumerate(self._row_values(report_type, row), 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        wb.save(path)
        return True, f"Excel report saved to {path}"
